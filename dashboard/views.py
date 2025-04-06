from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.db.models import Sum, Count
from django.utils import timezone
from products.models import Product, Category
from orders.models import Order, OrderItem
from core.models import UserProfile
from core.email_utils import send_order_status_update_email, send_order_shipped_email, send_order_delivered_email
from .models import SalesData
import openpyxl
from openpyxl.styles import Font
from datetime import datetime, timedelta
import csv
import io

@login_required
def dashboard(request):
    """Main dashboard view for sellers"""
    # Check if user is a seller
    if not request.user.profile.is_seller:
        messages.error(request, "You don't have permission to access the dashboard")
        return redirect('home')
    
    # Get dashboard statistics
    total_orders = Order.objects.count()
    total_products = Product.objects.count()
    total_customers = UserProfile.objects.filter(is_seller=False).count()
    total_sales = Order.objects.filter(status='delivered').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Get recent orders
    recent_orders = Order.objects.order_by('-created_at')[:5]
    
    # Get sales data for chart
    today = timezone.now().date()
    last_week = today - timedelta(days=7)
    
    # Create or update sales data
    update_sales_data()
    
    # Get sales data for the last 7 days
    sales_data = SalesData.objects.filter(date__gte=last_week).order_by('date')
    
    context = {
        'total_orders': total_orders,
        'total_products': total_products,
        'total_customers': total_customers,
        'total_sales': total_sales,
        'recent_orders': recent_orders,
        'sales_data': sales_data,
    }
    
    return render(request, 'dashboard/dashboard.html', context)

def update_sales_data():
    """Update sales data for reporting"""
    today = timezone.now().date()
    
    # Check if today's data already exists
    if not SalesData.objects.filter(date=today).exists():
        # Get today's orders
        today_orders = Order.objects.filter(created_at__date=today)
        total_sales = today_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        total_orders = today_orders.count()
        
        # Create sales data record
        SalesData.objects.create(
            date=today,
            total_sales=total_sales,
            total_orders=total_orders
        )

@login_required
def order_list(request):
    """View for listing all orders"""
    # Check if user is a seller
    if not request.user.profile.is_seller:
        messages.error(request, "You don't have permission to access the dashboard")
        return redirect('home')
    
    # Get filter parameters
    status = request.GET.get('status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Filter orders
    orders = Order.objects.all()
    
    if status:
        orders = orders.filter(status=status)
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__gte=date_from)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__lte=date_to)
        except ValueError:
            pass
    
    # Order by created_at (newest first)
    orders = orders.order_by('-created_at')
    
    context = {
        'orders': orders,
        'status_choices': Order.STATUS_CHOICES,
        'selected_status': status,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'dashboard/order_list.html', context)

@login_required
def order_detail(request, order_id):
    """View for displaying order details"""
    # Check if user is a seller
    if not request.user.profile.is_seller:
        messages.error(request, "You don't have permission to access the dashboard")
        return redirect('home')
    
    order = get_object_or_404(Order, id=order_id)
    
    context = {
        'order': order,
        'items': order.items.all(),
        'status_choices': Order.STATUS_CHOICES,
    }
    
    return render(request, 'dashboard/order_detail.html', context)

@login_required
def update_order_status(request, order_id):
    """View for updating order status"""
    # Check if user is a seller
    if not request.user.profile.is_seller:
        messages.error(request, "You don't have permission to access the dashboard")
        return redirect('home')
    
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        old_status = order.status
        status = request.POST.get('status')
        
        if status in dict(Order.STATUS_CHOICES).keys():
            order.status = status
            order.save()
            
            # Send email notification based on new status
            try:
                if status == 'processing':
                    send_order_status_update_email(order)
                elif status == 'shipped':
                    send_order_shipped_email(order)
                elif status == 'delivered':
                    send_order_delivered_email(order)
                elif status != old_status:  # For other status changes
                    send_order_status_update_email(order)
            except Exception as e:
                # Log the error but don't stop the process
                print(f"Error sending email: {e}")
            
            messages.success(request, f"Order #{order.id} status updated to {order.get_status_display()}")
        else:
            messages.error(request, "Invalid status")
    
    return redirect('dashboard_order_detail', order_id=order_id)

@login_required
def product_list(request):
    """View for listing all products"""
    # Check if user is a seller
    if not request.user.profile.is_seller:
        messages.error(request, "You don't have permission to access the dashboard")
        return redirect('home')
    
    # Get filter parameters
    category_id = request.GET.get('category')
    stock_status = request.GET.get('stock_status')
    
    # Filter products
    products = Product.objects.all()
    
    if category_id:
        products = products.filter(category_id=category_id)
    
    if stock_status == 'in_stock':
        products = products.filter(stock__gt=0)
    elif stock_status == 'out_of_stock':
        products = products.filter(stock=0)
    elif stock_status == 'low_stock':
        products = products.filter(stock__gt=0, stock__lte=10)
    
    # Order by name
    products = products.order_by('name')
    
    # Get all categories for filter dropdown
    categories = Category.objects.all()
    
    context = {
        'products': products,
        'categories': categories,
        'selected_category': category_id,
        'selected_stock_status': stock_status,
    }
    
    return render(request, 'dashboard/product_list.html', context)

@login_required
def add_product(request):
    """View for adding a new product"""
    # Check if user is a seller
    if not request.user.profile.is_seller:
        messages.error(request, "You don't have permission to access the dashboard")
        return redirect('home')
    
    # Get all categories
    categories = Category.objects.all()
    
    if request.method == 'POST':
        # Get form data
        name = request.POST.get('name')
        category_id = request.POST.get('category')
        description = request.POST.get('description')
        price = request.POST.get('price')
        stock = request.POST.get('stock')
        image = request.FILES.get('image')
        
        # Validate form data
        if not all([name, category_id, description, price, stock]):
            messages.error(request, "Please fill in all required fields")
            return render(request, 'dashboard/add_product.html', {'categories': categories})
        
        try:
            price = float(price)
            stock = int(stock)
            category = Category.objects.get(id=category_id)
        except (ValueError, Category.DoesNotExist):
            messages.error(request, "Invalid input data")
            return render(request, 'dashboard/add_product.html', {'categories': categories})
        
        # Create product
        product = Product(
            name=name,
            category=category,
            description=description,
            price=price,
            stock=stock
        )
        
        if image:
            product.image = image
        
        product.save()
        
        messages.success(request, f"Product '{name}' added successfully")
        return redirect('dashboard_products')
    
    return render(request, 'dashboard/add_product.html', {'categories': categories})

@login_required
def edit_product(request, product_id):
    """View for editing a product"""
    # Check if user is a seller
    if not request.user.profile.is_seller:
        messages.error(request, "You don't have permission to access the dashboard")
        return redirect('home')
    
    product = get_object_or_404(Product, id=product_id)
    categories = Category.objects.all()
    
    if request.method == 'POST':
        # Get form data
        name = request.POST.get('name')
        category_id = request.POST.get('category')
        description = request.POST.get('description')
        price = request.POST.get('price')
        stock = request.POST.get('stock')
        image = request.FILES.get('image')
        
        # Validate form data
        if not all([name, category_id, description, price, stock]):
            messages.error(request, "Please fill in all required fields")
            return render(request, 'dashboard/edit_product.html', {'product': product, 'categories': categories})
        
        try:
            price = float(price)
            stock = int(stock)
            category = Category.objects.get(id=category_id)
        except (ValueError, Category.DoesNotExist):
            messages.error(request, "Invalid input data")
            return render(request, 'dashboard/edit_product.html', {'product': product, 'categories': categories})
        
        # Update product
        product.name = name
        product.category = category
        product.description = description
        product.price = price
        product.stock = stock
        
        if image:
            product.image = image
        
        product.save()
        
        messages.success(request, f"Product '{name}' updated successfully")
        return redirect('dashboard_products')
    
    return render(request, 'dashboard/edit_product.html', {'product': product, 'categories': categories})

@login_required
def delete_product(request, product_id):
    """View for deleting a product"""
    # Check if user is a seller
    if not request.user.profile.is_seller:
        messages.error(request, "You don't have permission to access the dashboard")
        return redirect('home')
    
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        product_name = product.name
        product.delete()
        messages.success(request, f"Product '{product_name}' deleted successfully")
        return redirect('dashboard_products')
    
    return render(request, 'dashboard/delete_product.html', {'product': product})

@login_required
def customer_list(request):
    """View for listing all customers"""
    # Check if user is a seller
    if not request.user.profile.is_seller:
        messages.error(request, "You don't have permission to access the dashboard")
        return redirect('home')
    
    # Get all customers (users with is_seller=False)
    customers = UserProfile.objects.filter(is_seller=False).select_related('user')
    
    return render(request, 'dashboard/customer_list.html', {'customers': customers})

@login_required
def export_customers(request):
    """View for exporting customer data to Excel"""
    # Check if user is a seller
    if not request.user.profile.is_seller:
        messages.error(request, "You don't have permission to access the dashboard")
        return redirect('home')
    
    # Get all customers
    customers = UserProfile.objects.filter(is_seller=False).select_related('user')
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Add headers
    headers = ['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Phone', 'Address', 'Date Joined']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Add customer data
    for row_num, profile in enumerate(customers, 2):
        user = profile.user
        ws.cell(row=row_num, column=1).value = user.id
        ws.cell(row=row_num, column=2).value = user.username
        ws.cell(row=row_num, column=3).value = user.email
        ws.cell(row=row_num, column=4).value = user.first_name
        ws.cell(row=row_num, column=5).value = user.last_name
        ws.cell(row=row_num, column=6).value = profile.phone
        ws.cell(row=row_num, column=7).value = profile.address
        ws.cell(row=row_num, column=8).value = user.date_joined.strftime('%Y-%m-%d %H:%M:%S')
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=customers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    return response

@login_required
def sales_data(request):
    """View for displaying sales data"""
    # Check if user is a seller
    if not request.user.profile.is_seller:
        messages.error(request, "You don't have permission to access the dashboard")
        return redirect('home')
    
    # Get date range parameters
    period = request.GET.get('period', 'week')
    
    # Calculate date range based on period
    today = timezone.now().date()
    
    if period == 'week':
        start_date = today - timedelta(days=7)
        title = 'Last 7 Days'
    elif period == 'month':
        start_date = today - timedelta(days=30)
        title = 'Last 30 Days'
    elif period == 'year':
        start_date = today - timedelta(days=365)
        title = 'Last 365 Days'
    else:
        start_date = today - timedelta(days=7)
        title = 'Last 7 Days'
    
    # Get sales data for the selected period
    sales_data = SalesData.objects.filter(date__gte=start_date).order_by('date')
    
    # Calculate totals
    total_sales = sales_data.aggregate(Sum('total_sales'))['total_sales__sum'] or 0
    total_orders = sales_data.aggregate(Sum('total_orders'))['total_orders__sum'] or 0
    
    # Get top selling products
    top_products = OrderItem.objects.filter(
        order__created_at__date__gte=start_date
    ).values(
        'product__name'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_sales=Sum('price')
    ).order_by('-total_quantity')[:5]
    
    context = {
        'sales_data': sales_data,
        'total_sales': total_sales,
        'total_orders': total_orders,
        'top_products': top_products,
        'period': period,
        'title': title,
    }
    
    return render(request, 'dashboard/sales_data.html', context)
